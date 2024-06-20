import tkinter as tk
from tkinter import filedialog, ttk, simpledialog, messagebox
import os
import openpyxl
from openpyxl.styles import PatternFill, Color

class ExcelColorChanger:
    def __init__(self, master):
        self.master = master
        master.title("Excel Color Changer")

        self.notebook = ttk.Notebook(master)
        self.notebook.pack(expand=1, fill='both')

        # Tab for color input
        self.color_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.color_tab, text='Input Colors')
        
        self.color_prompt_label = tk.Label(self.color_tab, text="Masukkan kode warna asal dan warna baru:")
        self.color_prompt_label.pack(pady=10)
        
        self.original_color_label = tk.Label(self.color_tab, text="Warna Asal (hex, misalnya 'FF00FFFF' untuk cyan):")
        self.original_color_label.pack()
        self.original_color_entry = tk.Entry(self.color_tab)
        self.original_color_entry.pack()

        self.new_color_label = tk.Label(self.color_tab, text="Warna Baru (hex, misalnya '92D050' untuk hijau):")
        self.new_color_label.pack()
        self.new_color_entry = tk.Entry(self.color_tab)
        self.new_color_entry.pack()

        self.color_submit_button = tk.Button(self.color_tab, text="Submit", command=self.submit_colors)
        self.color_submit_button.pack(pady=10)

        # Tab for directory selection and processing
        self.process_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.process_tab, text='Change Fill Color')
        self.notebook.tab(self.process_tab, state='disabled')

        # Button to select directory
        self.select_button = tk.Button(self.process_tab, text="Pilih Direktori", command=self.select_directory)
        self.select_button.pack(pady=10)

        # Progress bar to display process status
        self.progress_bar = ttk.Progressbar(self.process_tab, orient='horizontal', mode='determinate')
        self.progress_bar.pack(fill='x', pady=10)

        # Label to display process status
        self.status_label = tk.Label(self.process_tab, text="")
        self.status_label.pack()

    def submit_colors(self):
        self.original_color = self.original_color_entry.get().strip()
        self.new_color = self.new_color_entry.get().strip()

        if not self.original_color or not self.new_color:
            messagebox.showerror("Error", "Warna asal dan warna baru harus dimasukkan!")
        else:
            try:
                # Validate color codes (they should be valid hex codes of length 8)
                if len(self.original_color) != 8 or len(self.new_color) != 8:
                    raise ValueError("Kode warna harus dalam format hex dengan panjang 8 karakter.")
                int(self.original_color, 16)
                int(self.new_color, 16)
                
                self.notebook.tab(self.process_tab, state='normal')
                self.notebook.select(self.process_tab)
            except ValueError as e:
                messagebox.showerror("Error", f"Kode warna tidak valid: {e}")

    def select_directory(self):
        # Open directory selection dialog
        directory = filedialog.askdirectory()
        if directory:
            self.status_label.config(text="Mengubah warna pada file Excel di direktori...")
            self.master.update()  # Update GUI to show status change
            self.change_cell_color_in_directory(directory)
            self.status_label.config(text="Selesai!")
            self.master.update()  # Update GUI to show status change

    def change_cell_color_in_directory(self, directory):
        # Find all Excel files in the directory
        excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

        # Set progress bar maximum value
        self.progress_bar['maximum'] = len(excel_files)
        self.progress_bar['value'] = 0

        for filename in excel_files:
            filepath = os.path.join(directory, filename)
            self.change_cell_color(filepath)
            self.progress_bar['value'] += 1
            self.master.update()  # Update progress bar

    def change_cell_color(self, filename):
        try:
            # Open Excel file
            workbook = openpyxl.load_workbook(filename)

            # Iterate through all sheets in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Find and change the color of cells with the original color
                for row in range(1, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.fill.start_color.index == self.original_color:
                            cell.fill = PatternFill(fgColor=Color(rgb=self.new_color), fill_type='solid')

            # Save changes to the Excel file
            workbook.save(filename)
        except Exception as e:
            messagebox.showerror("Error", f"Terjadi kesalahan saat memproses file {filename}: {e}")

root = tk.Tk()
app = ExcelColorChanger(root)
root.mainloop()
